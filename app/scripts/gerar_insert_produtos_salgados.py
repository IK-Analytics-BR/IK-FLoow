import openpyxl
from pathlib import Path

EXCEL_PATH = Path(r"C:\\Users\\arita\\Downloads\\02.02.2026.xlsx")

# Nomes exatos das categorias criadas no sistema
CAT_MATERIA_PRIMA = "Matéria Prima"
CAT_SEMI = "Semiacabado (Massa/Recheio/Caldo)"
CAT_PROD_FINAL = "Produto Acabado Salgado"
CAT_EMBALAGEM = "Embalagem"

UNIT_MAP = {
    "KG": "KG",
    "G": "G",
    "L": "L",
    "ML": "ML",
    "CX": "CX",
    "PCT": "PCT",
    "UN": "UN",
}


def sql_str(value: str) -> str:
    if value is None:
        return "NULL"
    s = str(value)
    s = s.replace("'", "''")
    return f"'{s}'"


def classify_category_from_mp(name: str):
    n = name.strip().upper()

    # Linhas que claramente não são produtos
    if n.startswith("28/"):
        return None
    if n.startswith("CUSTO"):
        return None
    if n.startswith("OBS.:"):
        return None
    if n.startswith("USOU "):
        return None
    if n == "PRODUTOS INTERMEDIARIOS":
        return None

    # Embalagens / etiquetagem
    if n.startswith("EMBALAGEM") or n.startswith("ETIQUETA") or n.startswith("ESPETO"):
        return CAT_EMBALAGEM

    # Produtos semiacabados (massas, recheios, caldos, coberturas, preparos prontos)
    semi_prefixes = [
        "MASSA ",
        "MASSA BOLINHO",
        "MASSA ESPETO",
        "MASSA MANDIOCA",
        "MASSA QUIBE",
        "MASSA TRIGO",
        "RECHEIO",
        "COBERTURA",
        "MANDIOCA COZIDA",
        "CARNE MOIDA COZIDA",
        "CARNE SECA DESFIADA",
        "CALDO CREME CEBOLA",
        "CALDO LIGANEX",
    ]
    if any(n.startswith(p) for p in semi_prefixes):
        return CAT_SEMI

    # Demais itens: matéria-prima
    return CAT_MATERIA_PRIMA


def choose_unit(pack_unit, base_unit, category_name: str) -> str:
    # Tenta primeiro a unidade base, depois a unidade da embalagem
    for raw in (base_unit, pack_unit):
        if isinstance(raw, str):
            u = raw.strip().upper()
            if u in UNIT_MAP:
                return UNIT_MAP[u]
    # Fallback por categoria
    if category_name == CAT_PROD_FINAL or category_name == CAT_EMBALAGEM:
        return "UN"
    # Padrão para matérias-primas e semiacabados
    return "KG"


def load_workbook():
    if not EXCEL_PATH.exists():
        raise SystemExit(f"Arquivo Excel não encontrado: {EXCEL_PATH}")
    return openpyxl.load_workbook(EXCEL_PATH, data_only=True)


def collect_mp_products(wb):
    if "LISTA MP" not in wb.sheetnames:
        return []
    ws = wb["LISTA MP"]
    products = []
    seen = set()

    # A partir da linha 3 começam os itens
    for row in ws.iter_rows(min_row=3, values_only=True):
        name = row[0]
        if not isinstance(name, str):
            continue
        name = name.strip()
        if not name or name in seen:
            continue

        category_name = classify_category_from_mp(name)
        if category_name is None:
            continue

        pack_unit = row[3]
        base_unit = row[5]
        unit = choose_unit(pack_unit, base_unit, category_name)

        description = f"{category_name} - {name}"

        products.append({
            "name": name,
            "category_name": category_name,
            "unit": unit,
            "description": description,
        })
        seen.add(name)

    return products


def collect_final_products(wb):
    final_names = set()
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            # PRODTO na primeira coluna
            if isinstance(row[0], str) and row[0].strip().upper() == "PRODUTO":
                if len(row) > 1 and isinstance(row[1], str) and row[1].strip():
                    final_names.add(row[1].strip())
            # PRODTO na segunda coluna
            if len(row) > 1 and isinstance(row[1], str) and row[1].strip().upper() == "PRODUTO":
                if len(row) > 2 and isinstance(row[2], str) and row[2].strip():
                    final_names.add(row[2].strip())

    products = []
    for name in sorted(final_names):
        description = f"Produto acabado salgado - {name}"
        products.append({
            "name": name,
            "category_name": CAT_PROD_FINAL,
            "unit": "UN",
            "description": description,
        })
    return products


def main():
    wb = load_workbook()

    mp_products = collect_mp_products(wb)
    final_products = collect_final_products(wb)

    all_products = mp_products + final_products
    # Acumular todas as linhas SQL em memória para poder gravar em arquivo UTF-8
    lines = []

    lines.append("USE supply_chain_system;")
    lines.append("")
    lines.append("INSERT INTO products (")
    lines.append("    name, description, unit_measure, category_id, category, ")
    lines.append("    price, cost_price, margin, max_discount, product_type, active")
    lines.append(") VALUES")

    values_lines = []
    for p in all_products:
        name = p["name"]
        desc = p["description"]
        unit = p["unit"]
        cat_name = p["category_name"]

        line = "    (" + ", ".join([
            sql_str(name),
            sql_str(desc),
            sql_str(unit),
            f"(SELECT id FROM product_categories WHERE name = {sql_str(cat_name)} LIMIT 1)",
            sql_str(cat_name),
            "0.00",  # price
            "0.00",  # cost_price
            "0.00",  # margin
            "0.00",  # max_discount
            sql_str("standalone"),
            "1",  # active
        ]) + ")"
        values_lines.append(line)

    # Junta todas as linhas com vírgula e finaliza com ponto-e-vírgula
    for i, line in enumerate(values_lines):
        suffix = "," if i < len(values_lines) - 1 else ";"
        lines.append(line + suffix)

    # Montar SQL completo
    sql = "\n".join(lines) + "\n"

    # Ainda imprimir no stdout (útil para debug / inspeção rápida)
    print(sql)

    # Gravar em arquivo UTF-8 para não perder acentuação no Windows
    output_path = Path(__file__).with_name("INSERT_PRODUTOS_SALGADOS.sql")
    output_path.write_text(sql, encoding="utf-8")
    print(f"\n[INFO] Arquivo SQL gerado em: {output_path}")


if __name__ == "__main__":
    main()
