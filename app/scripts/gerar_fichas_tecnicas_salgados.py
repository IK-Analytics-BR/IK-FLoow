import openpyxl
from pathlib import Path
from typing import List, Dict, Any, Optional

from openpyxl.utils.cell import column_index_from_string
import re

EXCEL_PATH = Path(r"C:\\Users\\arita\\Downloads\\02.02.2026.xlsx")

# Abas que contêm receitas de produtos finais salgados
RECIPE_SHEETS = {
    "70g",
    "20g",
    "Coxinha",
    "Salg.Assado",
    "Bolinho",
    "Quibe",
    "Risles",  # pode vir com problema de encoding no nome
    "Pastel",
}


def sql_str(value: Optional[str]) -> str:
    """Escapa string para uso em SQL."""
    if value is None:
        return "NULL"
    s = str(value)
    s = s.replace("'", "''")
    return f"'{s}'"


def normalize_unit(u: Optional[str]) -> Optional[str]:
    if not u:
        return None
    u = str(u).strip().upper()
    if u in {"UND", "UNDS", "UNID", "UNIDADE", "UN"}:
        return "UN"
    if u in {"PCTS", "PCT"}:
        return "PCT"
    if u in {"KG", "KGS"}:
        return "KG"
    if u in {"L", "LT", "LTS"}:
        return "L"
    if u in {"ML", "MLS"}:
        return "ML"
    if u.strip() == "M":
        return "M"
    return u


def classify_tipo_item(section: str, ingredient_name: str) -> str:
    """Define tipo_item para produto_template_itens.

    Por simplicidade inicial:
    - Tudo em EMPACOTAMENTO ou com nome iniciando com EMBALAGEM/ETIQUETA/ESPETO/RIBON => consumo_interno
    - Demais => materia_prima
    """
    sec = (section or "").strip().lower()
    n = ingredient_name.strip().upper()

    if sec == "empacotamento":
        return "consumo_interno"

    if n.startswith("EMBALAGEM") or n.startswith("ETIQUETA") or n.startswith("ESPETO") or n.startswith("RIBON"):
        return "consumo_interno"

    return "materia_prima"


def resolve_name(val_cell_value: Any, form_cell, wb_vals) -> Optional[str]:
    """Tenta obter um nome textual confiável para o ingrediente.

    1) Se o valor (data_only) for string e não começar com '#', usa ele.
    2) Se for erro/None, tenta olhar a fórmula e seguir referência simples do tipo =Sheet!A10
       para buscar o texto na outra aba (ex: Massas, Recheios).
    """
    if isinstance(val_cell_value, str):
        s = val_cell_value.strip()
        if s and not s.startswith("#"):
            return s

    if form_cell is None:
        return None

    fval = form_cell.value
    if not isinstance(fval, str) or not fval.startswith("="):
        return None

    formula = fval[1:]  # remove '='

    if "!" not in formula:
        return None

    sheet_ref, cell_ref = formula.split("!", 1)
    sheet_ref = sheet_ref.strip().strip("'")

    cell_ref = cell_ref.strip()
    m = re.match(r"\$?([A-Z]+)\$?(\d+)", cell_ref)
    if not m:
        return None

    col_letters, row_str = m.group(1), m.group(2)
    try:
        row_idx = int(row_str)
        col_idx = column_index_from_string(col_letters)
    except Exception:
        return None

    if sheet_ref not in wb_vals.sheetnames:
        return None

    ws_target = wb_vals[sheet_ref]
    try:
        target_cell = ws_target.cell(row=row_idx, column=col_idx)
    except Exception:
        return None

    tval = target_cell.value
    if isinstance(tval, str):
        s = tval.strip()
        if s and not s.startswith("#"):
            return s
    return None


def parse_sheet_templates(sheet_name: str, wb_vals, wb_form) -> List[Dict[str, Any]]:
    ws_vals = wb_vals[sheet_name]
    ws_form = wb_form[sheet_name]

    templates: List[Dict[str, Any]] = []
    current: Optional[Dict[str, Any]] = None
    current_section: Optional[str] = None  # 'massa', 'recheio', 'empacotamento'
    idx_per_unit: Optional[int] = None

    def flush_current():
        nonlocal current
        if current and current.get("itens"):
            templates.append(current)
        current = None

    for row_vals, row_forms in zip(ws_vals.iter_rows(values_only=True), ws_form.iter_rows(values_only=False)):
        # Garantir tamanho mínimo
        if not row_vals:
            continue

        first = row_vals[0]
        first_str = str(first).strip().upper() if first is not None else ""

        # Início de um novo produto
        if first_str == "PRODUTO" and row_vals[1]:
            flush_current()
            produto_nome = str(row_vals[1]).strip()
            current = {
                "sheet": sheet_name,
                "produto_nome": produto_nome,
                "rendimento": None,  # dict com info de rendimento, se houver
                "itens": [],  # lista de dicts
            }
            current_section = None
            idx_per_unit = None
            continue

        if current is None:
            continue

        # RENDIMENTO
        if first_str == "RENDIMENTO":
            # Ex: ('RENDIMENTO', 241.07, 'UNDS', 24, 24.10, 'PCTS', ...)
            unid_qty = row_vals[1]
            unid_label = row_vals[2]
            pct_qty = row_vals[3] if len(row_vals) > 3 else None
            pct_label = row_vals[5] if len(row_vals) > 5 else None
            current["rendimento"] = {
                "unidades": unid_qty,
                "unid_label": unid_label,
                "pcts": pct_qty,
                "pcts_label": pct_label,
            }
            continue

        # COMPOSIÇÃO - cabeçalho
        if first_str == "COMPOSIO":  # pode vir com encoding esquisito
            # Tentar achar coluna 'P/UND' ou 'P/UNID'
            idx_per_unit = None
            for idx, v in enumerate(row_vals):
                if isinstance(v, str):
                    vs = v.upper().replace("\u00a0", " ")
                    if "P/" in vs and "UND" in vs:
                        idx_per_unit = idx
                        break
            continue

        # Seção MASSA / RECHEIO / EMPACOTAMENTO
        if first_str.startswith("MASSA"):
            current_section = "massa"
            # Às vezes o cabeçalho de P/UND está nesta linha
            if idx_per_unit is None:
                for idx, v in enumerate(row_vals):
                    if isinstance(v, str):
                        vs = v.upper().replace("\u00a0", " ")
                        if "P/" in vs and "UND" in vs:
                            idx_per_unit = idx
                            break
            continue

        if first_str.startswith("RECHEIO"):
            current_section = "recheio"
            if idx_per_unit is None:
                for idx, v in enumerate(row_vals):
                    if isinstance(v, str):
                        vs = v.upper().replace("\u00a0", " ")
                        if "P/" in vs and "UND" in vs:
                            idx_per_unit = idx
                            break
            continue

        if first_str.startswith("EMPACOTAMENTO"):
            current_section = "empacotamento"
            # normalmente já temos idx_per_unit herdado
            continue

        # Linhas vazias / separadores - ignorar
        if not first_str:
            continue

        # Ignorar outras cabeceiras
        if first_str in {"LINHA", "PARAMETROS", "PESO IDEAL", "PESO MINIMO", "PESO MAXIMO"}:
            continue

        # Se não temos seção ou coluna de per-unit definida, não conseguimos processar item
        if not current_section or idx_per_unit is None or idx_per_unit >= len(row_vals):
            continue

        # Tentar obter quantidade por unidade a partir da coluna identificada
        qty_per_unit = row_vals[idx_per_unit]
        if qty_per_unit is None:
            continue
        if isinstance(qty_per_unit, str):
            # ignorar erros ou textos
            try:
                qty_per_unit = float(qty_per_unit.replace(",", "."))
            except Exception:
                continue

        try:
            qty_val = float(qty_per_unit)
        except Exception:
            continue

        if qty_val <= 0:
            continue

        # Nome do ingrediente: pode vir como string direta ou #REF! com fórmulas
        name_val = row_vals[0]
        form_cell = row_forms[0]
        ingredient_name = resolve_name(name_val, form_cell, wb_vals)
        if not ingredient_name:
            continue

        # Unidade do ingrediente (coluna 1, tipicamente)
        unit_val = row_vals[1] if len(row_vals) > 1 else None
        unit = normalize_unit(unit_val)

        tipo_item = classify_tipo_item(current_section, ingredient_name)

        current["itens"].append(
            {
                "nome": ingredient_name.strip(),
                "secao": current_section,
                "unidade": unit,
                "quantidade_por_unidade": qty_val,
                "tipo_item": tipo_item,
            }
        )

    # flush final
    flush_current()
    return templates


def build_sql(templates: List[Dict[str, Any]]) -> str:
    lines: List[str] = []
    lines.append("USE supply_chain_system;")
    lines.append("")

    # ================================
    # Mapear semiacabados (MASSA / RECHEIO) por produto
    # Cada produto final pode gerar até 2 semiacabados:
    # - MASSA - <produto>
    # - RECHEIO - <produto>
    # ================================
    semi_defs: Dict[tuple, Dict[str, Any]] = {}

    for t in templates:
        produto_nome = t["produto_nome"]
        sheet_name = t["sheet"]

        massa_itens = [item for item in t["itens"] if item["secao"] == "massa"]
        recheio_itens = [item for item in t["itens"] if item["secao"] == "recheio"]

        if massa_itens:
            key = ("massa", produto_nome)
            if key not in semi_defs:
                semi_defs[key] = {
                    "produto_nome": produto_nome,
                    "sheet": sheet_name,
                    "secao": "massa",
                    "itens": massa_itens,
                }

        if recheio_itens:
            key = ("recheio", produto_nome)
            if key not in semi_defs:
                semi_defs[key] = {
                    "produto_nome": produto_nome,
                    "sheet": sheet_name,
                    "secao": "recheio",
                    "itens": recheio_itens,
                }

    # ================================
    # INSERT em products para semiacabados (MASSA/RECHEIO)
    # ================================
    if semi_defs:
        lines.append("-- Produtos semiacabados (MASSA/RECHEIO) gerados automaticamente a partir das fichas dos salgados")
        for (secao, produto_nome), semi in semi_defs.items():
            semi_nome = f"{secao.upper()} - {produto_nome}"
            semi_nome_sql = sql_str(semi_nome)
            descricao_sql = sql_str(f"Semiacabado {secao.upper()} para {produto_nome}")
            categoria_sql = sql_str("Semiacabado (Massa/Recheio/Caldo)")

            lines.append(
                "INSERT INTO products ("  # campos mínimos necessários
                "internal_code, name, description, barcode, "
                "unit_measure, category_id, price, cost_price, margin, max_discount, active, category" \
                ")"
            )

            value_line = "SELECT " + ", ".join(
                [
                    "NULL",  # internal_code
                    semi_nome_sql,
                    descricao_sql,
                    "''",  # barcode
                    sql_str("KG"),  # unidade padrão para semiacabado
                    "NULL",  # category_id (pode ser ajustado depois via UI)
                    "0.00",  # price
                    "0.00",  # cost_price
                    "0.00",  # margin
                    "0.00",  # max_discount
                    "1",  # active
                    categoria_sql,
                ]
            ) + " WHERE NOT EXISTS (SELECT 1 FROM products WHERE name = " + semi_nome_sql + ");"

            lines.append(value_line)

        lines.append("")

    # ================================
    # INSERT em produto_templates_producao (produtos finais + semiacabados)
    # ================================
    lines.append("-- Templates de producao para produtos salgados e semiacabados")
    lines.append("INSERT INTO produto_templates_producao (")
    lines.append("    produto_id, versao, nome_template, custo_total_base,")
    lines.append("    tempo_producao_horas, ativo, observacoes, created_by")
    lines.append(") VALUES")

    template_values: List[str] = []
    used_products = set()

    # Templates dos produtos finais
    for t in templates:
        produto_nome = t["produto_nome"]
        if produto_nome in used_products:
            # Evitar duplicar template para o mesmo produto
            continue
        used_products.add(produto_nome)

        produto_expr = f"(SELECT id FROM products WHERE name = {sql_str(produto_nome)} LIMIT 1)"
        nome_template = f"Receita padrao - {produto_nome}"

        obs_parts = []
        rend = t.get("rendimento") or {}
        unid = rend.get("unidades")
        unid_label = rend.get("unid_label")
        pcts = rend.get("pcts")
        pcts_label = rend.get("pcts_label")
        if unid:
            obs_parts.append(f"Rendimento: {unid} {unid_label or ''}")
        if pcts:
            obs_parts.append(f"Equivalente a: {pcts} {pcts_label or ''}")
        obs_parts.append(f"Planilha: {t['sheet']}")
        observacoes = " | ".join(str(x) for x in obs_parts if x)

        value_line = "    (" + ", ".join(
            [
                produto_expr,
                "1",  # versao
                sql_str(nome_template),
                "0.00",  # custo_total_base (pode ser calculado depois)
                "NULL",  # tempo_producao_horas
                "1",  # ativo
                sql_str(observacoes),
                "NULL",  # created_by
            ]
        ) + ")"
        template_values.append(value_line)

    # Templates dos semiacabados (MASSA / RECHEIO)
    for (secao, produto_nome), semi in semi_defs.items():
        semi_nome = f"{secao.upper()} - {produto_nome}"
        semi_prod_expr = f"(SELECT id FROM products WHERE name = {sql_str(semi_nome)} LIMIT 1)"
        nome_template = f"{secao.capitalize()} para {produto_nome}"

        obs_parts = [
            f"Semiacabado {secao} derivado de {produto_nome}",
            f"Planilha: {semi['sheet']}",
        ]
        observacoes = " | ".join(str(x) for x in obs_parts if x)

        value_line = "    (" + ", ".join(
            [
                semi_prod_expr,
                "1",  # versao
                sql_str(nome_template),
                "0.00",  # custo_total_base
                "NULL",  # tempo_producao_horas
                "1",  # ativo
                sql_str(observacoes),
                "NULL",  # created_by
            ]
        ) + ")"
        template_values.append(value_line)

    if not template_values:
        lines.append("    (NULL, 1, 'SEM_TEMPLATES', 0.00, NULL, 0, 'Nenhum template gerado', NULL);")
        return "\n".join(lines) + "\n"

    for i, vline in enumerate(template_values):
        suffix = "," if i < len(template_values) - 1 else ";"
        lines.append(vline + suffix)

    lines.append("")

    # ================================
    # INSERT em produto_template_itens
    # ================================
    lines.append("-- Itens das fichas tecnicas (quantidades POR UNIDADE produzida)")
    lines.append("INSERT INTO produto_template_itens (")
    lines.append("    template_id, tipo_item, produto_id, descricao,")
    lines.append("    quantidade, unidade_medida, custo_unitario_base, custo_total_base, observacoes")
    lines.append(") VALUES")

    item_values: List[str] = []

    # Itens dos templates dos produtos finais (mantém estrutura atual)
    for t in templates:
        produto_nome = t["produto_nome"]
        template_id_expr = (
            "(SELECT t.id FROM produto_templates_producao t "
            "JOIN products p ON t.produto_id = p.id "
            f"WHERE p.name = {sql_str(produto_nome)} AND t.versao = 1 LIMIT 1)"
        )

        for item in t["itens"]:
            ing_nome = item["nome"]
            tipo_item = item["tipo_item"]
            unidade = item["unidade"]
            qtd = item["quantidade_por_unidade"]
            secao = item["secao"]

            produto_ing_expr = (
                f"(SELECT id FROM products WHERE name = {sql_str(ing_nome)} LIMIT 1)"
            )

            descricao_item = f"{secao.upper()} - {ing_nome}"

            value_line = "    (" + ", ".join(
                [
                    template_id_expr,
                    sql_str(tipo_item),
                    produto_ing_expr,
                    sql_str(descricao_item),
                    f"{qtd:.6f}",
                    sql_str(unidade),
                    "NULL",  # custo_unitario_base
                    "NULL",  # custo_total_base
                    sql_str(f"Importado de {t['sheet']}")
                ]
            ) + ")"
            item_values.append(value_line)

    # Itens dos templates dos semiacabados (MASSA / RECHEIO)
    for (secao, produto_nome), semi in semi_defs.items():
        semi_nome = f"{secao.upper()} - {produto_nome}"
        template_id_expr = (
            "(SELECT t.id FROM produto_templates_producao t "
            "JOIN products p ON t.produto_id = p.id "
            f"WHERE p.name = {sql_str(semi_nome)} AND t.versao = 1 LIMIT 1)"
        )

        for item in semi["itens"]:
            ing_nome = item["nome"]
            tipo_item = item["tipo_item"]
            unidade = item["unidade"]
            qtd = item["quantidade_por_unidade"]

            produto_ing_expr = (
                f"(SELECT id FROM products WHERE name = {sql_str(ing_nome)} LIMIT 1)"
            )

            descricao_item = f"{secao.upper()} - {ing_nome}"

            value_line = "    (" + ", ".join(
                [
                    template_id_expr,
                    sql_str(tipo_item),
                    produto_ing_expr,
                    sql_str(descricao_item),
                    f"{qtd:.6f}",
                    sql_str(unidade),
                    "NULL",  # custo_unitario_base
                    "NULL",  # custo_total_base
                    sql_str(f"Semiacabado {secao} derivado de {produto_nome}")
                ]
            ) + ")"
            item_values.append(value_line)

    if not item_values:
        lines.append("    (NULL, 'materia_prima', NULL, 'SEM_ITENS', 0.0, NULL, NULL, NULL, NULL);")
        return "\n".join(lines) + "\n"

    for i, vline in enumerate(item_values):
        suffix = "," if i < len(item_values) - 1 else ";"
        lines.append(vline + suffix)

    return "\n".join(lines) + "\n"


def main() -> None:
    if not EXCEL_PATH.exists():
        raise SystemExit(f"Arquivo Excel não encontrado: {EXCEL_PATH}")

    # Carregar duas vezes: uma com valores calculados, outra com fórmulas para resolver nomes
    wb_vals = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    wb_form = openpyxl.load_workbook(EXCEL_PATH, data_only=False)

    all_templates: List[Dict[str, Any]] = []

    for sheet_name in wb_vals.sheetnames:
        if sheet_name not in RECIPE_SHEETS:
            continue
        try:
            templates = parse_sheet_templates(sheet_name, wb_vals, wb_form)
            all_templates.extend(templates)
        except Exception as e:
            print(f"[WARN] Falha ao processar aba {sheet_name}: {e}")

    if not all_templates:
        print("Nenhuma ficha técnica encontrada nas abas selecionadas.")
        return

    sql = build_sql(all_templates)

    # Imprimir no stdout para inspeção
    print(sql)

    # Gravar em arquivo .sql UTF-8 ao lado do script
    output_path = Path(__file__).with_name("INSERT_FICHAS_TECNICAS_SALGADOS.sql")
    output_path.write_text(sql, encoding="utf-8")
    print(f"\n[INFO] Arquivo SQL de fichas técnicas gerado em: {output_path}")


if __name__ == "__main__":
    main()
